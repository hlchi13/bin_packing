from itertools import chain, combinations
import math
from threading import Timer

from pysat.formula import CNF
from pysat.solvers import Solver

import matplotlib.pyplot as plt
import timeit
from typing import List
from pysat.solvers import Glucose3, Solver
from prettytable import PrettyTable
from threading import Timer
import datetime
import pandas as pd
import os
import sys
import time
from openpyxl import load_workbook
from openpyxl import Workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

class TimeoutException(Exception): pass
# Initialize the CNF formula
n_items = 0
W, H = 0, 0
items = []

#read file
def read_input():
    global W, H, items, n_items
    n_items = int(input().split()[0])
    W, H = map(int, input().split())
    for i in range(n_items):
        list = input().split() 
        items.append([int(list[0]), int(list[1])])
    
def positive_range(end):
    if (end < 0):
        return []
    return range(end)
def compare_item_by_longer_side(item):
    return max(item[0], item[1])
def BPP_result(rectangles, W, H, max_bins, n_items):
# Define the variables
    height = H
    width = W
    cnf = CNF()
    variables = {}
    counter = 1

    # x_i_j = 1 if item i is placed in bin j
    for i in range(n_items):
        for i2 in range(max_bins):
            variables[f"x{i + 1},{i2 + 1}"] = counter 
            counter += 1

    for i in range(len(rectangles)):
        for i2 in range(len(rectangles)):
            if i != i2:
                variables[f"lr{i + 1},{i2 + 1}"] = counter  # lri,rj
                counter += 1
                variables[f"ud{i + 1},{i2 + 1}"] = counter  # uri,rj
                counter += 1
        for e in range(width):
            variables[f"px{i + 1},{e}"] = counter  # pxi,e
            counter += 1
        for f in range(height):
            variables[f"py{i + 1},{f}"] = counter  # pyi,f
            counter += 1

    # Rotated variables
    for i in range(len(rectangles)):
        variables[f"r{i + 1}"] = counter
        counter += 1

    # Exactly one bin for each item
    for i in range(n_items):
        cnf.append([variables[f"x{i + 1},{j + 1}"] for j in range(max_bins)])
        for j1 in range(max_bins):
            for j2 in range(j1 + 1, max_bins):
                cnf.append([-variables[f"x{i + 1},{j1 + 1}"], -variables[f"x{i + 1},{j2 + 1}"]])

    # Add the 2-literal axiom clauses
    for i in range(len(rectangles)):
        for e in range(width - 1):  # -1 because we're using e+1 in the clause
            cnf.append([-variables[f"px{i + 1},{e}"],
                        variables[f"px{i + 1},{e + 1}"]])
        for f in range(height - 1):  # -1 because we're using f+1 in the clause
            cnf.append([-variables[f"py{i + 1},{f}"],
                        variables[f"py{i + 1},{f + 1}"]])

    # Add non-overlapping constraints
    def non_overlapping(rotated, i1, i2, h1, h2, v1, v2, j):
        if not rotated:
            i1_width = rectangles[i1][0]
            i1_height = rectangles[i1][1]
            i2_width = rectangles[i2][0]
            i2_height = rectangles[i2][1]
            i1_rotation = variables[f"r{i1 + 1}"]
            i2_rotation = variables[f"r{i2 + 1}"]
        else:
            i1_width = rectangles[i1][1]
            i1_height = rectangles[i1][0]
            i2_width = rectangles[i2][1]
            i2_height = rectangles[i2][0]
            i1_rotation = -variables[f"r{i1 + 1}"]
            i2_rotation = -variables[f"r{i2 + 1}"]

        # Square symmertry breaking, if i is square than it cannot be rotated
        if i1_width == i2_height and rotated:
            i_square = True
            cnf.append([-variables[f"r{i1 + 1}"]])
        else:
            i_square = False

        if i2_width == i2_height and rotated:
            j_square = True
            cnf.append([-variables[f"r{i2 + 1}"]])
        else:
            j_square = False
        bin_cnf = [-variables[f"x{i1 + 1},{j + 1}"], -variables[f"x{i2 + 1},{j + 1}"]]
        # lri,j v lrj,i v udi,j v udj,i
        four_literal = []
        if h1: four_literal.append(variables[f"lr{i1 + 1},{i2 + 1}"])
        if h2: four_literal.append(variables[f"lr{i2 + 1},{i1 + 1}"])
        if v1: four_literal.append(variables[f"ud{i1 + 1},{i2 + 1}"])
        if v2: four_literal.append(variables[f"ud{i2 + 1},{i1 + 1}"])

        cnf.append(four_literal + [i1_rotation] + bin_cnf)
        cnf.append(four_literal + [i2_rotation] + bin_cnf)

        # ¬lri, j ∨ ¬pxj, e
        if h1:
            for e in range(min(width, i1_width)):
                    cnf.append(bin_cnf + [i1_rotation,
                                -variables[f"lr{i1 + 1},{i2 + 1}"],
                                -variables[f"px{i2 + 1},{e}"]])
        # ¬lrj,i ∨ ¬pxi,e
        if h2:
            for e in range(min(width, i2_width)):
                    cnf.append(bin_cnf + [i2_rotation,
                                -variables[f"lr{i2 + 1},{i1 + 1}"],
                                -variables[f"px{i1 + 1},{e}"]])
        # ¬udi,j ∨ ¬pyj,f
        if v1:
            for f in range(min(height, i1_height)):
                    cnf.append(bin_cnf + [i1_rotation,
                                -variables[f"ud{i1 + 1},{i2 + 1}"],
                                -variables[f"py{i2 + 1},{f}"]])
        # ¬udj, i ∨ ¬pyi, f,
        if v2:
            for f in range(min(height, i2_height)):
                    cnf.append(bin_cnf + [i2_rotation,
                                -variables[f"ud{i2 + 1},{i1 + 1}"],
                                -variables[f"py{i1 + 1},{f}"]])

        for e in positive_range(width - i1_width):
            # ¬lri,j ∨ ¬pxj,e+wi ∨ pxi,e
            if h1:
                    cnf.append(bin_cnf + [i1_rotation,
                                -variables[f"lr{i1 + 1},{i2 + 1}"],
                                variables[f"px{i1 + 1},{e}"],
                                -variables[f"px{i2 + 1},{e + i1_width}"]])

        for e in positive_range(width - i2_width):
            # ¬lrj,i ∨ ¬pxi,e+wj ∨ pxj,e
            cnf.append(bin_cnf + [i2_rotation,
                                -variables[f"lr{i2 + 1},{i1 + 1}"],
                                variables[f"px{i2 + 1},{e}"],
                                -variables[f"px{i1 + 1},{e + i2_width}"]])

        for f in positive_range(height - i1_height):
            # udi,j ∨ ¬pyj,f+hi ∨ pxi,e
            if v1:
                    cnf.append(bin_cnf + [i1_rotation,
                                -variables[f"ud{i1 + 1},{i2 + 1}"],
                                variables[f"py{i1 + 1},{f}"],
                                -variables[f"py{i2 + 1},{f + i1_height}"]])
        for f in positive_range(height - i2_height):
            # ¬udj,i ∨ ¬pyi,f+hj ∨ pxj,f
            if v2:
                    cnf.append(bin_cnf + [i2_rotation,
                                -variables[f"ud{i2 + 1},{i1 + 1}"],
                                variables[f"py{i2 + 1},{f}"],
                                -variables[f"py{i1 + 1},{f + i2_height}"]])
    for j in range(max_bins):
        for i in range(len(rectangles)):
            for i2 in range(i + 1, len(rectangles)):
            #  #Large-rectangles horizontal
                if min(rectangles[i][0], rectangles[i][1]) + min(rectangles[i2][0], rectangles[i2][1]) > width:
                    non_overlapping(False, i, i2, False, False, True, True, j)
                    non_overlapping(True, i, i2, False, False, True, True, j)
                # Large rectangles vertical
                elif min(rectangles[i][0], rectangles[i][1]) + min(rectangles[i2][0], rectangles[i2][1]) > height:
                    non_overlapping(False, i, i2, True, True, False, False, j)
                    non_overlapping(True, i, i2, True, True, False, False, j)

                # Same rectangle and is a square
                elif rectangles[i] == rectangles[i2]:
                    if rectangles[i][0] == rectangles[i][1]:
                        cnf.append([-variables[f"r{i + 1}"]])
                        cnf.append([-variables[f"r{i2 + 1}"]])
                        non_overlapping(False,i ,i2, True, False, True, True, j)
                    else:
                        non_overlapping(False,i ,i2, True, False, True, True, j)
                        non_overlapping(True,i ,i2, True, False, True, True, j)
            # #normal rectangles
                else:
                    non_overlapping(False, i, i2, True, True, True, True, j)
                    non_overlapping(True, i, i2, True, True, True, True, j)

    # Domain encoding to ensure every rectangle stays inside strip's boundary
    for i2 in range(max_bins):
        for i in range(len(rectangles)):
            if rectangles[i][0] > width: #if rectangle[i]'s width larger than strip's width, it has to be rotated
                cnf.append([variables[f"r{i + 1}"]])
            else:
                cnf.append([variables[f"r{i + 1}"],
                                    variables[f"px{i + 1},{width - rectangles[i][0]}"]])
       
            if rectangles[i][1] > height:
                cnf.append([variables[f"r{i + 1}"]])
            else:
                cnf.append([variables[f"r{i + 1}"],
                            variables[f"py{i + 1},{height - rectangles[i][1]}"]])

            # Rotated
            if rectangles[i][1] > width:
                cnf.append([-variables[f"r{i + 1}"]])
            else:
                
                cnf.append([-variables[f"r{i + 1}"],
                                    variables[f"px{i + 1},{width - rectangles[i][1]}"]])
            if rectangles[i][0] > height:
                cnf.append([-variables[f"r{i + 1}"]])
            else:
                cnf.append([-variables[f"r{i + 1}"],
                                variables[f"py{i + 1},{height - rectangles[i][0]}"]])
    start = timeit.default_timer()
    with Solver(name="mc") as solver: #add all cnf to solver
        solver.append_formula(cnf)

        if solver.solve():
            pos = [[0 for i in range(2)] for j in range(len(rectangles))]
            rotation = []
            model = solver.get_model()
            print("SAT")
            result = {}
            bin_used = []
            solver_time = format(timeit.default_timer() - start, ".3f")
            for var in model:
                if var > 0:
                    result[list(variables.keys())[list(variables.values()).index(var)]] = True
                else:
                    result[list(variables.keys())[list(variables.values()).index(-var)]] = False
            for i in range(len(rectangles)):
                rotation.append(result[f"r{i + 1}"])
                for e in range(width - 1):
                    if result[f"px{i + 1},{e}"] == False and result[f"px{i + 1},{e + 1}"] == True:
                        pos[i][0] = e + 1
                    if e == 0 and result[f"px{i + 1},{e}"] == True:
                        pos[i][0] = 0
                for f in range(height - 1):
                    if result[f"py{i + 1},{f}"] == False and result[f"py{i + 1},{f + 1}"] == True:
                        pos[i][1] = f + 1
                    if f == 0 and result[f"py{i + 1},{f}"] == True:
                        pos[i][1] = 0
            for i2 in range(max_bins):
                bin_used.append([i for i in range(n_items) if result[f"x{i + 1},{i2 + 1}"] == True])
            return(["sat", bin_used, pos, rotation, solver_time, len(variables), len(cnf.clauses)])

        else:
            print("unsat")
            return("unsat")
def interrupt(solver):
    solver.interrupt()
  
def BPP(W, H, items, n_items, max_bins):
    items_area = [i[0] * i[1] for i in items]
    bin_area = W * H
    lower_bound = math.ceil(sum(items_area) / bin_area)
    print(sum(items_area))
    for k in range(lower_bound, max_bins + 1):
        print(f"Trying with {k} bins")
        result = BPP_result(items, W, H, max_bins=k, n_items=n_items)
        if result[0] == "sat":
            print(f"Solution found with {k} bins")
            return result[1:]
        
    return None
    
	
def print_solution(bpp_result):
    result_dict = {}
    if bpp_result is None:
        print("No solution found")
        return
    else:
        bins = bpp_result[0]
        pos = bpp_result[1]
        rotation = bpp_result[2]
        solver_time = bpp_result[3]
        num_variables = bpp_result[4]
        num_clauses = bpp_result[5]
        print(pos)
        print(rotation)
        for i in range(len(bins)):
            print("Bin", i + 1, "contains items", [(j + 1) for j in bins[i]])
            for j in bins[i]:
                if rotation[j]:
                    print("Rotated item", j + 1, items[j], "at position", pos[j])
                else:
                    print("Item", j + 1, items[j], "at position", pos[j])
            # display_solution((W, H), [items[j] for j in bins[i]], [pos[j] for j in bins[i]], [rotation[j] for j in bins[i]])
        print("--------------------")
        print("Solution found with", len(bins), "bins")
        print("Solver time:", solver_time)
        print("Number of variables:", num_variables)
        print("Number of clauses:", num_clauses)
        result_dict = {
            "Type": "not using OPP",
            "Data": os.path.basename(sys.argv[1]),
            "Number of items": n_items,
            "Minimize Bin": len(bins),  
            "Solver time": solver_time, 
            "Number of variables": num_variables, 
            "Number of clauses": num_clauses}
    write_to_xlsx(result_dict)

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path = 'out/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)
    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

    print(f"Result added to Excel file: {os.path.abspath(excel_file_path)}\n")

def display_solution(strip, rectangles, pos_circuits, rotation):
    # define Matplotlib figure and axis
    fig, ax = plt.subplots()
    ax = plt.gca()
    plt.title(strip)
    if len(pos_circuits) > 0:
        for i in range(len(rectangles)):
            rect = plt.Rectangle(pos_circuits[i],
                                 rectangles[i][0] if not rotation[i] else rectangles[i][1],
                                 rectangles[i][1] if not rotation[i] else rectangles[i][0],
                                 edgecolor="#333")
            ax.add_patch(rect)

    ax.set_xlim(0, strip[0])
    ax.set_ylim(0, strip[1] + 1)
    ax.set_xticks(range(strip[0] + 1))
    ax.set_yticks(range(strip[1] + 1))
    ax.set_xlabel('width')
    ax.set_ylabel('height')
    # display plot
    plt.show()

def solve():
    # read input file
    global W, H, items, n_items
    if len(sys.argv) < 2:
        print("Error: No file name provided.")
        return
    
    with open(sys.argv[1], 'r') as f:
        sys.stdin = f
        
        start = time.time()
        
        read_input()
        print(W, H)

        bpp_result = BPP(W, H, items, n_items, n_items)
        stop = time.time()
        print_solution(bpp_result)
        print("Time:", stop - start)

if __name__ == "__main__":
    solve()